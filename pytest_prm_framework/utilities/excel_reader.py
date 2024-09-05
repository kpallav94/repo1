import openpyxl
def get_list_from_excel(excel_name,sheet_name):
   excel=openpyxl.load_workbook(excel_name)
   sheet=excel[sheet_name]
   values=[]
   for row in range(1,sheet.max_row+1):
     nested_values=[]
     for column in range(1,sheet.max_column+1):
        data=sheet.cell(row,column)
        nested_values.append(data.value)
     values.append(nested_values)
   return values


def add_value_to_excel(excel_name,sheet_name):
    excel=openpyxl.load_workbook(excel_name)
    sheet=excel[sheet_name]
    sheet.cell(10,1).value="new value"
    excel.save(excel_name)

add_value_to_excel("../creds.xlsx","login_creds")